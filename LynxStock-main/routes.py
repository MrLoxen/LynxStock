# routes.py
from flask import Blueprint, render_template, request, jsonify, send_file, current_app
from models import db, Proveedor, Producto, Cliente, Venta, DetalleVenta, Credito, Abono, TasaBCV, Compra, DetalleCompra, CompraCredito, PagoProveedor, DeudaGeneral, PagoDeudaGeneral
from datetime import datetime, date, timedelta
from io import BytesIO
import openpyxl
import os
from scraper_bcv import obtener_tasa_bcv  

api = Blueprint('api', __name__)

# ---------- Páginas principales ----------
@api.route('/')
def inicio():
    return render_template('inicio.html')

@api.route('/inventario')
def inventario_page():
    return render_template('inventario.html')

@api.route('/proveedores')
def proveedores_page():
    return render_template('proveedores.html')

@api.route('/clientes')
def clientes_page():
    return render_template('clientes.html')

@api.route('/caja')
def caja_page():
    return render_template('caja.html')

@api.route('/creditos')
def creditos_page():
    return render_template('creditos.html')

@api.route('/reportes')
def reportes_page():
    return render_template('reportes.html')

@api.route('/compras')
def compras_page():
    return render_template('compras.html')

@api.route('/cuentas-pagar')
def cuentas_pagar_page():
    return render_template('cuentas_pagar.html')

# ---------- Endpoint para la tasa BCV ----------
@api.route('/api/tasa-bcv', methods=['GET'])
def obtener_tasa_actual():
    ultima = TasaBCV.query.order_by(TasaBCV.id.desc()).first()
    if ultima:
        return jsonify({
            'tasa': ultima.tasa,
            'fecha': ultima.fecha.isoformat()
        })
    return jsonify({'tasa': 0, 'fecha': None}), 404

# ---------- API: Proveedores ----------
@api.route('/api/proveedores', methods=['GET', 'POST'])
def manejar_proveedores():
    if request.method == 'GET':
        proveedores = Proveedor.query.all()
        return jsonify([{'id': p.id, 'nombre': p.nombre, 'rif': p.rif, 'telefono': p.telefono} for p in proveedores])
    elif request.method == 'POST':
        data = request.get_json()
        prov = Proveedor(nombre=data['nombre'], rif=data.get('rif'), telefono=data.get('telefono'))
        db.session.add(prov)
        db.session.commit()
        return jsonify({'id': prov.id}), 201

@api.route('/api/proveedores/<int:id>', methods=['PUT', 'DELETE'])
def modificar_proveedor(id):
    prov = Proveedor.query.get_or_404(id)
    if request.method == 'PUT':
        data = request.get_json()
        prov.nombre = data.get('nombre', prov.nombre)
        prov.rif = data.get('rif', prov.rif)
        prov.telefono = data.get('telefono', prov.telefono)
        db.session.commit()
        return jsonify({'mensaje': 'Proveedor actualizado'})
    elif request.method == 'DELETE':
        db.session.delete(prov)
        db.session.commit()
        return jsonify({'mensaje': 'Proveedor eliminado'})

# ---------- API: Productos ----------
@api.route('/api/productos', methods=['GET', 'POST'])
def manejar_productos():
    if request.method == 'GET':
        productos = Producto.query.all()
        resultado = []
        for pr in productos:
            resultado.append({
                'id': pr.id,
                'codigo': pr.codigo,
                'descripcion': pr.descripcion,
                'proveedor_id': pr.proveedor_id,
                'proveedor_nombre': pr.proveedor.nombre if pr.proveedor else '',
                'costo': pr.costo,
                'precio': pr.precio,
                'stock': pr.stock,
                'stock_minimo': pr.stock_minimo,
                'ubicacion': pr.ubicacion
            })
        return jsonify(resultado)
    elif request.method == 'POST':
        data = request.get_json()
        prod = Producto(
            codigo=data['codigo'],
            descripcion=data['descripcion'],
            proveedor_id=data.get('proveedor_id'),
            costo=data.get('costo', 0),
            precio=data.get('precio', 0),
            stock=data.get('stock', 0),
            stock_minimo=data.get('stock_minimo', 0),
            ubicacion=data.get('ubicacion', '')
        )
        db.session.add(prod)
        db.session.commit()
        return jsonify({'id': prod.id}), 201

@api.route('/api/productos/<int:id>', methods=['PUT', 'DELETE'])
def modificar_producto(id):
    prod = Producto.query.get_or_404(id)
    if request.method == 'PUT':
        data = request.get_json()
        prod.codigo = data.get('codigo', prod.codigo)
        prod.descripcion = data.get('descripcion', prod.descripcion)
        prod.proveedor_id = data.get('proveedor_id', prod.proveedor_id)
        prod.costo = data.get('costo', prod.costo)
        prod.precio = data.get('precio', prod.precio)
        prod.stock = data.get('stock', prod.stock)
        prod.stock_minimo = data.get('stock_minimo', prod.stock_minimo)
        prod.ubicacion = data.get('ubicacion', prod.ubicacion)
        db.session.commit()
        return jsonify({'mensaje': 'Producto actualizado'})
    elif request.method == 'DELETE':
        db.session.delete(prod)
        db.session.commit()
        return jsonify({'mensaje': 'Producto eliminado'})

# ---------- API: Clientes ----------
@api.route('/api/clientes', methods=['GET', 'POST'])
def manejar_clientes():
    if request.method == 'GET':
        clientes = Cliente.query.all()
        return jsonify([{'id': c.id, 'nombre': c.nombre, 'cedula': c.cedula,
                         'telefono': c.telefono, 'saldo_pendiente': c.saldo_pendiente} for c in clientes])
    elif request.method == 'POST':
        data = request.get_json()
        if Cliente.query.filter_by(cedula=data['cedula']).first():
            return jsonify({'error': 'Cédula ya registrada'}), 400
        cli = Cliente(nombre=data['nombre'], cedula=data['cedula'], telefono=data.get('telefono'))
        db.session.add(cli)
        db.session.commit()
        return jsonify({'id': cli.id}), 201

@api.route('/api/clientes/<int:id>', methods=['PUT', 'DELETE'])
def modificar_cliente(id):
    cli = Cliente.query.get_or_404(id)
    if request.method == 'PUT':
        data = request.get_json()
        cli.nombre = data.get('nombre', cli.nombre)
        cli.telefono = data.get('telefono', cli.telefono)
        db.session.commit()
        return jsonify({'mensaje': 'Cliente actualizado'})
    elif request.method == 'DELETE':
        db.session.delete(cli)
        db.session.commit()
        return jsonify({'mensaje': 'Cliente eliminado'})

# ---------- API: Ventas ----------
@api.route('/api/ventas', methods=['POST'])
def crear_venta():
    data = request.get_json()
    tipo = data['tipo']
    cliente_id = data.get('cliente_id')
    productos_vendidos = data['productos']

    tasa = 0.0
    ultima_tasa = TasaBCV.query.order_by(TasaBCV.id.desc()).first()
    if not ultima_tasa:
        return jsonify({'error': 'No hay tasa BCV registrada. Ejecute primero la aplicación.'}), 500
    tasa = ultima_tasa.tasa

    total_usd = 0.0
    total_bs = 0.0
    detalles = []
    for item in productos_vendidos:
        prod = Producto.query.get(item['id'])
        if not prod:
            return jsonify({'error': f'Producto con id {item["id"]} no encontrado'}), 400
        cantidad = float(item['cantidad'])
        if prod.stock < cantidad:
            return jsonify({'error': f'Stock insuficiente de {prod.descripcion}'}), 400

        precio_usd = prod.precio
        subtotal_usd = cantidad * precio_usd
        subtotal_bs = subtotal_usd * tasa

        total_usd += subtotal_usd
        total_bs += subtotal_bs
        detalles.append({
            'producto': prod,
            'cantidad': cantidad,
            'precio_unit': precio_usd
        })

    if tipo == 'credito' and not cliente_id:
        return jsonify({'error': 'Debe seleccionar un cliente para venta a crédito'}), 400

    venta = Venta(
        tipo=tipo,
        cliente_id=cliente_id,
        total_bs=total_bs,
        total_usd=total_usd,
        tasa_bcv=tasa
    )
    db.session.add(venta)
    db.session.flush()

    for det in detalles:
        det_venta = DetalleVenta(
            venta_id=venta.id,
            producto_id=det['producto'].id,
            cantidad=det['cantidad'],
            precio_unitario=det['precio_unit']
        )
        db.session.add(det_venta)
        det['producto'].stock = round(det['producto'].stock - det['cantidad'], 2)

    if tipo == 'credito':
        fecha_venc = data.get('fecha_vencimiento')
        if fecha_venc:
            fecha_venc = datetime.strptime(fecha_venc, '%Y-%m-%d').date()
        else:
            fecha_venc = date.today() + timedelta(days=30)
        credito = Credito(
            venta_id=venta.id,
            cliente_id=cliente_id,
            monto_total=total_bs,
            saldo_restante=total_bs,
            fecha_vencimiento=fecha_venc
        )
        db.session.add(credito)
        cliente = Cliente.query.get(cliente_id)
        if cliente:
            cliente.saldo_pendiente += total_bs

    db.session.commit()
    return jsonify({
        'mensaje': 'Venta registrada',
        'venta_id': venta.id,
        'total_usd': total_usd,
        'total_bs': total_bs,
        'tasa': tasa
    }), 201

# ---------- API: Créditos ----------
@api.route('/api/creditos', methods=['GET'])
def listar_creditos():
    estado = request.args.get('estado', 'pendiente')
    query = Credito.query
    if estado != 'todos':
        query = query.filter_by(estado=estado)
    creditos = query.order_by(Credito.fecha_vencimiento).all()
    resultado = []
    for c in creditos:
        cliente_nombre = ''
        if c.cliente_id:
            cliente = Cliente.query.get(c.cliente_id)
            if cliente:
                cliente_nombre = cliente.nombre
        resultado.append({
            'id': c.id,
            'venta_id': c.venta_id,
            'cliente_nombre': cliente_nombre,
            'monto_total': c.monto_total,
            'saldo_restante': c.saldo_restante,
            'fecha_vencimiento': c.fecha_vencimiento.isoformat() if c.fecha_vencimiento else '',
            'estado': c.estado,
            'dias_restantes': (c.fecha_vencimiento - date.today()).days if c.fecha_vencimiento else None
        })
    return jsonify(resultado)

@api.route('/api/creditos/<int:id>/abonar', methods=['POST'])
def abonar_credito(id):
    credito = Credito.query.get_or_404(id)
    data = request.get_json()
    monto = round(float(data['monto']), 2)

    if monto <= 0:
        return jsonify({'error': 'Monto debe ser positivo'}), 400
    if monto > credito.saldo_restante + 0.001:
        return jsonify({'error': 'El abono supera el saldo pendiente'}), 400

    abono = Abono(credito_id=id, monto=monto)
    db.session.add(abono)
    credito.saldo_restante = round(credito.saldo_restante - monto, 2)
    if credito.saldo_restante <= 0:
        credito.saldo_restante = 0.0
        credito.estado = 'pagado'

    cliente = Cliente.query.get(credito.cliente_id)
    if cliente:
        cliente.saldo_pendiente = round(cliente.saldo_pendiente - monto, 2)
        if cliente.saldo_pendiente < 0:
            cliente.saldo_pendiente = 0.0

    db.session.commit()
    return jsonify({'mensaje': 'Abono registrado', 'saldo_restante': credito.saldo_restante})

# ---------- Endpoint para contar productos con stock bajo ----------
@api.route('/api/productos/stock-bajo', methods=['GET'])
def contar_stock_bajo():
    cantidad = Producto.query.filter(Producto.stock <= Producto.stock_minimo).count()
    return jsonify({'count': cantidad})

# ---------- COMPRAS / ENTRADAS DE INVENTARIO ----------
@api.route('/api/compras', methods=['GET', 'POST'])
def manejar_compras():
    if request.method == 'GET':
        compras = Compra.query.order_by(Compra.fecha.desc()).all()
        resultado = []
        for c in compras:
            proveedor_nombre = ''
            if c.proveedor_id:
                prov = Proveedor.query.get(c.proveedor_id)
                if prov:
                    proveedor_nombre = prov.nombre
            resultado.append({
                'id': c.id,
                'fecha': c.fecha.strftime('%Y-%m-%d %H:%M'),
                'proveedor': proveedor_nombre,
                'total': c.total,
                'observaciones': c.observaciones
            })
        return jsonify(resultado)
    elif request.method == 'POST':
        data = request.get_json()
        proveedor_id = data.get('proveedor_id')
        observaciones = data.get('observaciones', '')
        productos_comprados = data['productos']

        total_compra = 0.0
        detalles = []
        for item in productos_comprados:
            prod = Producto.query.get(item['id'])
            if not prod:
                return jsonify({'error': f'Producto con id {item["id"]} no encontrado'}), 400
            cantidad = float(item['cantidad'])
            costo_unitario = float(item.get('costo_unitario', 0))
            subtotal = cantidad * costo_unitario
            total_compra += subtotal
            detalles.append({
                'producto': prod,
                'cantidad': cantidad,
                'costo_unitario': costo_unitario
            })

        tipo = data.get('tipo', 'contado')
        compra = Compra(
            proveedor_id=proveedor_id,
            total=total_compra,
            observaciones=observaciones
        )
        db.session.add(compra)
        db.session.flush()

        if tipo == 'credito':
            fecha_venc = data.get('fecha_vencimiento')
            if fecha_venc:
                fecha_venc = datetime.strptime(fecha_venc, '%Y-%m-%d').date()
            else:
                fecha_venc = date.today() + timedelta(days=30)

            credito = CompraCredito(
                compra_id=compra.id,
                proveedor_id=proveedor_id,
                monto_total=total_compra,
                saldo_restante=total_compra,
                fecha_vencimiento=fecha_venc
            )
            db.session.add(credito)
            proveedor = Proveedor.query.get(proveedor_id)
            if proveedor:
                proveedor.saldo_pagar = round(proveedor.saldo_pagar + total_compra, 2)

        for det in detalles:
            det_compra = DetalleCompra(
                compra_id=compra.id,
                producto_id=det['producto'].id,
                cantidad=det['cantidad'],
                costo_unitario=det['costo_unitario']
            )
            db.session.add(det_compra)
            det['producto'].stock += det['cantidad']
            det['producto'].costo = det['costo_unitario']

        db.session.commit()
        return jsonify({'mensaje': 'Compra registrada', 'compra_id': compra.id}), 201

@api.route('/api/compras/<int:id>', methods=['GET'])
def detalle_compra(id):
    compra = Compra.query.get_or_404(id)
    proveedor_nombre = ''
    if compra.proveedor_id:
        prov = Proveedor.query.get(compra.proveedor_id)
        if prov:
            proveedor_nombre = prov.nombre
    detalles = []
    for det in compra.detalles:
        prod = Producto.query.get(det.producto_id)
        detalles.append({
            'producto': prod.descripcion if prod else 'Producto eliminado',
            'codigo': prod.codigo if prod else '',
            'cantidad': det.cantidad,
            'costo_unitario': det.costo_unitario,
            'subtotal': det.cantidad * det.costo_unitario
        })
    return jsonify({
        'id': compra.id,
        'fecha': compra.fecha.strftime('%Y-%m-%d %H:%M'),
        'proveedor': proveedor_nombre,
        'total': compra.total,
        'observaciones': compra.observaciones,
        'detalles': detalles
    })

# ---------- Reporte diario de ventas ----------
@api.route('/api/ventas/hoy', methods=['GET'])
def ventas_hoy():
    hoy = date.today()
    inicio = datetime(hoy.year, hoy.month, hoy.day, 0, 0, 0)
    fin = datetime(hoy.year, hoy.month, hoy.day, 23, 59, 59)

    # --- Ventas del día (contado + créditos pagados) ---
    ventas_dia = Venta.query.filter(Venta.fecha >= inicio, Venta.fecha <= fin).order_by(Venta.fecha.desc()).all()

    ventas = []
    for v in ventas_dia:
        if v.tipo == 'contado':
            ventas.append(v)
        elif v.tipo == 'credito':
            credito = Credito.query.filter_by(venta_id=v.id).first()
            if credito and credito.estado == 'pagado':
                ventas.append(v)

    total_usd = 0.0
    total_bs = 0.0
    lista_ventas = []

    for v in ventas:
        total_usd += v.total_usd
        total_bs += v.total_bs

        cliente_nombre = None
        if v.cliente_id:
            cliente = Cliente.query.get(v.cliente_id)
            if cliente:
                cliente_nombre = cliente.nombre

        productos = []
        for det in v.detalles:
            prod = Producto.query.get(det.producto_id)
            productos.append({
                'descripcion': prod.descripcion if prod else 'Producto eliminado',
                'cantidad': det.cantidad
            })

        lista_ventas.append({
            'id': v.id,
            'hora': v.fecha.strftime('%H:%M:%S'),
            'tipo': v.tipo,
            'cliente': cliente_nombre,
            'total_usd': v.total_usd,
            'total_bs': v.total_bs,
            'tasa_bcv': v.tasa_bcv,
            'productos': productos
        })

    # --- Abonos del día ---
    abonos_dia = Abono.query.filter(Abono.fecha >= inicio, Abono.fecha <= fin).order_by(Abono.fecha.desc()).all()

    total_abonos_bs = 0.0
    lista_abonos = []
    for a in abonos_dia:
        total_abonos_bs += a.monto
        # Obtener nombre del cliente a través del crédito
        cliente_nombre = ''
        if a.credito and a.credito.cliente_id:
            cliente = Cliente.query.get(a.credito.cliente_id)
            if cliente:
                cliente_nombre = cliente.nombre
        lista_abonos.append({
            'id': a.id,
            'hora': a.fecha.strftime('%H:%M:%S'),
            'cliente': cliente_nombre,
            'monto': a.monto
        })

    return jsonify({
        'fecha': hoy.isoformat(),
        'resumen': {
            'cantidad_ventas': len(ventas),
            'total_ventas_usd': total_usd,
            'total_ventas_bs': total_bs,
            'total_abonos_bs': total_abonos_bs,
            'cantidad_abonos': len(abonos_dia)
        },
        'ventas': lista_ventas,
        'abonos': lista_abonos
    })

# ---------- Funciones Excel ----------
@api.route('/api/inventario/exportar')
def exportar_inventario():
    productos = Producto.query.all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario"
    ws.append(['Código', 'Descripción', 'Proveedor', 'Costo', 'Precio', 'Stock', 'Stock Mínimo', 'Ubicación'])
    for p in productos:
        ws.append([
            p.codigo,
            p.descripcion,
            p.proveedor.nombre if p.proveedor else '',
            p.costo,
            p.precio,
            p.stock,
            p.stock_minimo,
            p.ubicacion
        ])
    archivo = BytesIO()
    wb.save(archivo)
    archivo.seek(0)
    return send_file(archivo, download_name='inventario.xlsx', as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@api.route('/api/inventario/importar', methods=['POST'])
def importar_inventario():
    if 'file' not in request.files:
        return jsonify({'error': 'No se envió archivo'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Archivo sin nombre'}), 400

    filename = file.filename
    ruta = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)
    file.save(ruta)

    try:
        wb = openpyxl.load_workbook(ruta)
        ws = wb.active
        filas_importadas = 0
        errores = []

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if len(row) < 2:
                continue

            codigo = str(row[0]).strip() if row[0] else None
            if not codigo:
                continue

            descripcion = str(row[1]) if row[1] else ''
            proveedor_nombre = str(row[2]) if len(row) > 2 and row[2] else ''

            def a_numero(valor, nombre_columna, fila_num):
                if valor is None or str(valor).strip() == '':
                    return 0.0
                try:
                    return float(valor)
                except ValueError:
                    errores.append(f"Fila {fila_num}, columna '{nombre_columna}': '{valor}' no es un número. Se asumió 0.")
                    return 0.0

            costo = a_numero(row[3] if len(row) > 3 else None, 'Costo', idx)
            precio = a_numero(row[4] if len(row) > 4 else None, 'Precio', idx)
            stock = a_numero(row[5] if len(row) > 5 else None, 'Stock', idx)
            stock_min = a_numero(row[6] if len(row) > 6 else None, 'Stock Mínimo', idx)
            ubicacion = str(row[7]) if len(row) > 7 and row[7] else ''

            proveedor_id = None
            if proveedor_nombre:
                prov = Proveedor.query.filter_by(nombre=proveedor_nombre).first()
                if not prov:
                    prov = Proveedor(nombre=proveedor_nombre)
                    db.session.add(prov)
                    db.session.flush()
                proveedor_id = prov.id

            producto = Producto.query.filter_by(codigo=codigo).first()
            if producto:
                producto.descripcion = descripcion
                producto.proveedor_id = proveedor_id or producto.proveedor_id
                producto.costo = costo
                producto.precio = precio
                producto.stock = stock
                producto.stock_minimo = stock_min
                producto.ubicacion = ubicacion
            else:
                nuevo = Producto(
                    codigo=codigo,
                    descripcion=descripcion,
                    proveedor_id=proveedor_id,
                    costo=costo,
                    precio=precio,
                    stock=stock,
                    stock_minimo=stock_min,
                    ubicacion=ubicacion
                )
                db.session.add(nuevo)
            filas_importadas += 1

        db.session.commit()
        os.remove(ruta)

        mensaje = f'Se importaron/actualizaron {filas_importadas} productos.'
        if errores:
            mensaje += f' Se encontraron {len(errores)} advertencias (revisa la consola o los logs).'
            for e in errores:
                print(e)
        return jsonify({'mensaje': mensaje})
    except Exception as e:
        return jsonify({'error': f'Error al procesar el archivo: {str(e)}'}), 500

@api.route('/api/respaldo/completo')
def respaldo_completo():
    wb = openpyxl.Workbook()
    # Productos
    ws1 = wb.active
    ws1.title = "Productos"
    ws1.append(['ID', 'Código', 'Descripción', 'Proveedor', 'Costo', 'Precio', 'Stock'])
    for p in Producto.query.all():
        ws1.append([p.id, p.codigo, p.descripcion,
                    p.proveedor.nombre if p.proveedor else '', p.costo, p.precio, p.stock])

    # Clientes
    ws2 = wb.create_sheet("Clientes")
    ws2.append(['ID', 'Nombre', 'Cédula', 'Teléfono', 'Saldo Pendiente'])
    for c in Cliente.query.all():
        ws2.append([c.id, c.nombre, c.cedula, c.telefono, c.saldo_pendiente])

    # Ventas
    ws3 = wb.create_sheet("Ventas")
    ws3.append(['ID', 'Fecha', 'Tipo', 'Cliente', 'Total USD', 'Total Bs', 'Tasa BCV'])
    for v in Venta.query.all():
        cliente_nombre = 'Mostrador'
        if v.cliente_id:
            cliente = Cliente.query.get(v.cliente_id)
            if cliente:
                cliente_nombre = cliente.nombre
        ws3.append([v.id, v.fecha.strftime('%Y-%m-%d %H:%M'), v.tipo, cliente_nombre,
                    v.total_usd, v.total_bs, v.tasa_bcv])

    # Créditos Clientes
    ws4 = wb.create_sheet("Creditos Clientes")
    ws4.append(['ID', 'Cliente', 'Monto Total', 'Saldo Restante', 'Vencimiento', 'Estado'])
    for c in Credito.query.all():
        cliente_nombre = ''
        if c.cliente_id:
            cliente = Cliente.query.get(c.cliente_id)
            if cliente:
                cliente_nombre = cliente.nombre
        ws4.append([c.id, cliente_nombre, c.monto_total, c.saldo_restante,
                    c.fecha_vencimiento.isoformat() if c.fecha_vencimiento else '', c.estado])

    archivo = BytesIO()
    wb.save(archivo)
    archivo.seek(0)
    return send_file(archivo, download_name='respaldo_completo.xlsx', as_attachment=True)

# ---------- CRÉDITOS DE PROVEEDORES ----------
@api.route('/api/compras/creditos', methods=['GET'])
def listar_creditos_compras():
    estado = request.args.get('estado', 'pendiente')
    query = CompraCredito.query
    if estado != 'todos':
        query = query.filter_by(estado=estado)
    creditos = query.order_by(CompraCredito.fecha_vencimiento).all()
    resultado = []
    for c in creditos:
        proveedor_nombre = ''
        prov = Proveedor.query.get(c.proveedor_id)
        if prov:
            proveedor_nombre = prov.nombre
        resultado.append({
            'id': c.id,
            'compra_id': c.compra_id,
            'proveedor_nombre': proveedor_nombre,
            'monto_total': c.monto_total,
            'saldo_restante': c.saldo_restante,
            'fecha_vencimiento': c.fecha_vencimiento.isoformat() if c.fecha_vencimiento else '',
            'estado': c.estado,
            'dias_restantes': (c.fecha_vencimiento - date.today()).days if c.fecha_vencimiento else None
        })
    return jsonify(resultado)

@api.route('/api/compras/creditos/<int:id>/pagar', methods=['POST'])
def pagar_credito_compra(id):
    credito = CompraCredito.query.get_or_404(id)
    data = request.get_json()
    monto = round(float(data['monto']), 2)
    if monto <= 0:
        return jsonify({'error': 'Monto debe ser positivo'}), 400
    if monto > credito.saldo_restante + 0.001:
        return jsonify({'error': 'El pago supera el saldo pendiente'}), 400

    pago = PagoProveedor(compra_credito_id=id, monto=monto)
    db.session.add(pago)
    credito.saldo_restante = round(credito.saldo_restante - monto, 2)
    if credito.saldo_restante <= 0:
        credito.saldo_restante = 0.0
        credito.estado = 'pagado'

    proveedor = Proveedor.query.get(credito.proveedor_id)
    if proveedor:
        proveedor.saldo_pagar = round(proveedor.saldo_pagar - monto, 2)
        if proveedor.saldo_pagar < 0:
            proveedor.saldo_pagar = 0.0

    db.session.commit()
    return jsonify({'mensaje': 'Pago registrado', 'saldo_restante': credito.saldo_restante})

# ---------- DEUDAS GENERALES (gastos) ----------
@api.route('/api/deudas', methods=['GET', 'POST'])
def manejar_deudas():
    if request.method == 'GET':
        deudas = DeudaGeneral.query.order_by(DeudaGeneral.fecha_vencimiento).all()
        return jsonify([{
            'id': d.id,
            'descripcion': d.descripcion,
            'monto_total': d.monto_total,
            'saldo_restante': d.saldo_restante,
            'fecha_vencimiento': d.fecha_vencimiento.isoformat() if d.fecha_vencimiento else '',
            'estado': d.estado,
            'dias_restantes': (d.fecha_vencimiento - date.today()).days if d.fecha_vencimiento else None
        } for d in deudas])
    elif request.method == 'POST':
        data = request.get_json()
        fecha_venc = data.get('fecha_vencimiento')
        if fecha_venc:
            fecha_venc = datetime.strptime(fecha_venc, '%Y-%m-%d').date()
        else:
            fecha_venc = date.today() + timedelta(days=30)
        deuda = DeudaGeneral(
            descripcion=data['descripcion'],
            monto_total=data['monto_total'],
            saldo_restante=data['monto_total'],
            fecha_vencimiento=fecha_venc
        )
        db.session.add(deuda)
        db.session.commit()
        return jsonify({'mensaje': 'Deuda registrada', 'id': deuda.id}), 201

@api.route('/api/deudas/<int:id>/pagar', methods=['POST'])
def pagar_deuda(id):
    deuda = DeudaGeneral.query.get_or_404(id)
    data = request.get_json()
    monto = round(float(data['monto']), 2)
    if monto <= 0:
        return jsonify({'error': 'Monto debe ser positivo'}), 400
    if monto > deuda.saldo_restante + 0.001:
        return jsonify({'error': 'El pago supera el saldo pendiente'}), 400

    pago = PagoDeudaGeneral(deuda_id=id, monto=monto)
    db.session.add(pago)
    deuda.saldo_restante = round(deuda.saldo_restante - monto, 2)
    if deuda.saldo_restante <= 0:
        deuda.saldo_restante = 0.0
        deuda.estado = 'pagado'
    db.session.commit()
    return jsonify({'mensaje': 'Pago registrado', 'saldo_restante': deuda.saldo_restante})

# ---------- Búsqueda rápida de productos ----------
@api.route('/api/productos/buscar', methods=['GET'])
def buscar_productos():
    q = request.args.get('q', '').strip()
    if not q:
        return jsonify([])
    productos = Producto.query.filter(
        db.or_(
            Producto.codigo.ilike(f'%{q}%'),
            Producto.descripcion.ilike(f'%{q}%')
        )
    ).limit(10).all()
    return jsonify([{
        'id': p.id,
        'codigo': p.codigo,
        'descripcion': p.descripcion,
        'precio': p.precio,
        'stock': p.stock
    } for p in productos])