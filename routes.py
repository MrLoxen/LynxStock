# routes.py
from flask import Blueprint, render_template, request, jsonify, send_file, current_app, redirect, url_for
from flask_login import login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
from models import db, Proveedor, Producto, Cliente, Venta, DetalleVenta, Credito, Abono, TasaBCV, Compra, DetalleCompra, CompraCredito, PagoProveedor, DeudaGeneral, PagoDeudaGeneral, Devolucion, DetalleDevolucion, Usuario, Auditoria
from datetime import datetime, date, timedelta
from io import BytesIO
import openpyxl
import os
from scraper_bcv import obtener_tasa_bcv  

api = Blueprint('api', __name__)

# ---------- DECORADOR PARA ROLES ----------
def role_required(*roles):
    def wrapper(fn):
        @wraps(fn)
        def decorated_view(*args, **kwargs):
            if not current_user.is_authenticated or current_user.rol not in roles:
                return "Acceso denegado. Tu rol no tiene permisos para ver esta página.", 403
            return fn(*args, **kwargs)
        return decorated_view
    return wrapper

# ---------- SISTEMA DE LOGIN Y LOGOUT ----------
@api.route('/login', methods=['GET', 'POST'])
def login_page():
    if request.method == 'POST':
        data = request.get_json()
        username = data.get('username')
        password = data.get('password')
        
        user = Usuario.query.filter_by(username=username).first()
        if user and check_password_hash(user.password_hash, password):
            login_user(user)
            
            registrar_auditoria('LOGIN', 'SEGURIDAD', f'El usuario "{user.username}" ingresó al sistema con el rol {user.rol}.')
            
            return jsonify({'mensaje': 'Login exitoso', 'rol': user.rol}), 200
        return jsonify({'error': 'Usuario o contraseña incorrectos'}), 401
        
    return render_template('login.html')

@api.route('/logout')
@login_required
def logout():
    
    registrar_auditoria('LOGOUT', 'SEGURIDAD', f'El usuario "{current_user.username}" cerró sesión y salió del sistema.')
    
    logout_user()
    return redirect(url_for('api.login_page'))

@api.route('/crear-admin-secreto')
def crear_admin():
    if not Usuario.query.filter_by(username='admin').first():
        nuevo_admin = Usuario(
            username='admin', 
            password_hash=generate_password_hash('123456'), 
            rol='admin'
        )
        db.session.add(nuevo_admin)
        db.session.commit()
        return "Admin creado exitosamente. Ya puedes ir a /login"
    return "El admin ya existe."

# ---------- FUNCIÓN MAESTRA Y VISTAS DE AUDITORÍA ----------
def registrar_auditoria(accion, modulo, detalles):
    try:
        usuario_id = current_user.id if current_user.is_authenticated else None
        registro = Auditoria(
            usuario_id=usuario_id,
            accion=accion,
            modulo=modulo,
            detalles=detalles
        )
        db.session.add(registro)
        db.session.commit()
    except Exception as e:
        print(f"Error guardando auditoría: {e}")
        db.session.rollback()

@api.route('/auditoria')
@login_required
@role_required('admin')
def auditoria_page():
    return render_template('auditoria.html')

@api.route('/api/auditoria', methods=['GET'])
@login_required
@role_required('admin')
def api_auditoria():
    registros = Auditoria.query.order_by(Auditoria.fecha.desc()).limit(500).all()
    resultado = []
    for r in registros:
        resultado.append({
            'id': r.id,
            'fecha': r.fecha.strftime('%Y-%m-%d %H:%M:%S'),
            'usuario': r.usuario.username if r.usuario else 'Sistema',
            'accion': r.accion,
            'modulo': r.modulo,
            'detalles': r.detalles
        })
    return jsonify(resultado)

# ---------- GESTIÓN DE USUARIOS (SOLO ADMIN) ----------
@api.route('/usuarios')
@login_required
@role_required('admin')
def usuarios_page():
    return render_template('usuarios.html')

@api.route('/api/usuarios', methods=['GET', 'POST'])
@login_required
@role_required('admin')
def api_usuarios():
    if request.method == 'GET':
        usuarios = Usuario.query.all()
        return jsonify([{'id': u.id, 'username': u.username, 'rol': u.rol} for u in usuarios])
        
    elif request.method == 'POST':
        data = request.get_json()
        username = data.get('username')
        password = data.get('password')
        rol = data.get('rol')

        if Usuario.query.filter_by(username=username).first():
            return jsonify({'error': 'Ya existe un usuario con ese nombre. Elige otro.'}), 400

        nuevo_usuario = Usuario(
            username=username,
            password_hash=generate_password_hash(password),
            rol=rol
        )
        db.session.add(nuevo_usuario)
        db.session.commit()
        return jsonify({'mensaje': 'Usuario creado'}), 201

@api.route('/api/usuarios/<int:id>', methods=['DELETE'])
@login_required
@role_required('admin')
def eliminar_usuario(id):
    if current_user.id == id:
        return jsonify({'error': '¡No puedes eliminar tu propia cuenta mientras la usas!'}), 400
        
    usuario = Usuario.query.get_or_404(id)
    db.session.delete(usuario)
    db.session.commit()
    return jsonify({'mensaje': 'Usuario eliminado'})


# ---------- Páginas principales (PROTEGIDAS) ----------
@api.route('/')
@login_required
def inicio():
    return render_template('inicio.html')

@api.route('/inventario')
@login_required
@role_required('admin', 'gerente')
def inventario_page():
    return render_template('inventario.html')

@api.route('/proveedores')
@login_required
@role_required('admin', 'gerente')
def proveedores_page():
    return render_template('proveedores.html')

@api.route('/clientes')
@login_required
@role_required('admin', 'gerente', 'cajero')
def clientes_page():
    return render_template('clientes.html')

@api.route('/caja')
@login_required
@role_required('admin', 'gerente', 'cajero')
def caja_page():
    return render_template('caja.html')

@api.route('/creditos')
@login_required
@role_required('admin', 'gerente')
def creditos_page():
    return render_template('creditos.html')

@api.route('/reportes')
@login_required
@role_required('admin', 'gerente')
def reportes_page():
    return render_template('reportes.html')

@api.route('/compras')
@login_required
@role_required('admin', 'gerente')
def compras_page():
    return render_template('compras.html')

@api.route('/cuentas-pagar')
@login_required
@role_required('admin', 'gerente')
def cuentas_pagar_page():
    return render_template('cuentas_pagar.html')

@api.route('/devoluciones')
@login_required
@role_required('admin', 'gerente', 'cajero')
def devoluciones_page():
    return render_template('devoluciones.html')

# ---------- Endpoint para la tasa BCV ----------
@api.route('/api/tasa-bcv', methods=['GET'])
def obtener_tasa_actual():
    ultima = TasaBCV.query.order_by(TasaBCV.id.desc()).first()
    if ultima:
        return jsonify({
            'tasa': ultima.tasa,
            'fecha': ultima.fecha.isoformat()
        })
    return jsonify({'tasa': 0, 'fecha': None}), 404

# ---------- API: Proveedores ----------
@api.route('/api/proveedores', methods=['GET', 'POST'])
def manejar_proveedores():
    if request.method == 'GET':
        proveedores = Proveedor.query.all()
        return jsonify([{'id': p.id, 'nombre': p.nombre, 'rif': p.rif, 'telefono': p.telefono} for p in proveedores])
    elif request.method == 'POST':
        data = request.get_json()
        prov = Proveedor(nombre=data['nombre'], rif=data.get('rif'), telefono=data.get('telefono'))
        db.session.add(prov)
        db.session.commit()
        
        registrar_auditoria('CREAR', 'PROVEEDORES', f'Registrado nuevo proveedor: {prov.nombre} (RIF: {prov.rif or "N/A"}).')
        
        return jsonify({'id': prov.id}), 201

@api.route('/api/proveedores/<int:id>', methods=['PUT', 'DELETE'])
def modificar_proveedor(id):
    prov = Proveedor.query.get_or_404(id)
    if request.method == 'PUT':
        data = request.get_json()
        prov.nombre = data.get('nombre', prov.nombre)
        prov.rif = data.get('rif', prov.rif)
        prov.telefono = data.get('telefono', prov.telefono)
        db.session.commit()
        
        registrar_auditoria('EDITAR', 'PROVEEDORES', f'Modificados datos del proveedor: "{prov.nombre}" (ID: {prov.id}).')
        
        return jsonify({'mensaje': 'Proveedor actualizado'})
        
    elif request.method == 'DELETE':
        nombre_temporal = prov.nombre  
        db.session.delete(prov)
        db.session.commit()
        
        registrar_auditoria('ELIMINAR', 'PROVEEDORES', f'Eliminado permanentemente el proveedor: "{nombre_temporal}" (ID: {id}).')
        
        return jsonify({'mensaje': 'Proveedor eliminado'})

# ---------- API: Productos ----------
@api.route('/api/productos', methods=['GET', 'POST'])
def manejar_productos():
    if request.method == 'GET':
        productos = Producto.query.all()
        resultado = []
        for pr in productos:
            resultado.append({
                'id': pr.id,
                'codigo': pr.codigo,
                'descripcion': pr.descripcion,
                'proveedor_id': pr.proveedor_id,
                'proveedor_nombre': pr.proveedor.nombre if pr.proveedor else '',
                'costo': pr.costo,
                'precio': pr.precio,
                'stock': pr.stock,
                'stock_minimo': pr.stock_minimo,
                'ubicacion': pr.ubicacion
            })
        return jsonify(resultado)
    elif request.method == 'POST':
        data = request.get_json()
        prod = Producto(
            codigo=data['codigo'],
            descripcion=data['descripcion'],
            proveedor_id=data.get('proveedor_id'),
            costo=data.get('costo', 0),
            precio=data.get('precio', 0),
            stock=data.get('stock', 0),
            stock_minimo=data.get('stock_minimo', 0),
            ubicacion=data.get('ubicacion', '')
        )
        db.session.add(prod)
        db.session.commit()
        
        registrar_auditoria('CREAR', 'INVENTARIO', f'Creado nuevo producto: "{prod.descripcion}" (Código: {prod.codigo}).')
        
        return jsonify({'id': prod.id}), 201

@api.route('/api/productos/<int:id>', methods=['PUT', 'DELETE'])
def modificar_producto(id):
    prod = Producto.query.get_or_404(id)
    if request.method == 'PUT':
        data = request.get_json()
        prod.codigo = data.get('codigo', prod.codigo)
        prod.descripcion = data.get('descripcion', prod.descripcion)
        prod.proveedor_id = data.get('proveedor_id', prod.proveedor_id)
        prod.costo = data.get('costo', prod.costo)
        prod.precio = data.get('precio', prod.precio)
        prod.stock = data.get('stock', prod.stock)
        prod.stock_minimo = data.get('stock_minimo', prod.stock_minimo)
        prod.ubicacion = data.get('ubicacion', prod.ubicacion)
        db.session.commit()
        
        registrar_auditoria('EDITAR', 'INVENTARIO', f'Modificado el producto: "{prod.descripcion}" (Código: {prod.codigo}).')
        
        return jsonify({'mensaje': 'Producto actualizado'})
        
    elif request.method == 'DELETE':
        codigo_temp = prod.codigo
        desc_temp = prod.descripcion
        db.session.delete(prod)
        db.session.commit()
        
        registrar_auditoria('ELIMINAR', 'INVENTARIO', f'Eliminado del catálogo el producto: "{desc_temp}" (Código: {codigo_temp}).')
        
        return jsonify({'mensaje': 'Producto eliminado'})

# ---------- API: Clientes ----------
@api.route('/api/clientes', methods=['GET', 'POST'])
def manejar_clientes():
    if request.method == 'GET':
        clientes = Cliente.query.all()
        # Aquí añadimos 'direccion': c.direccion para que el front-end lo lea
        return jsonify([{'id': c.id, 'nombre': c.nombre, 'cedula': c.cedula,
                         'telefono': c.telefono, 'direccion': c.direccion, 'saldo_pendiente': c.saldo_pendiente} for c in clientes])
    elif request.method == 'POST':
        data = request.get_json()
        if Cliente.query.filter_by(cedula=data['cedula']).first():
            return jsonify({'error': 'Cédula ya registrada'}), 400
        cli = Cliente(
            nombre=data['nombre'], 
            cedula=data['cedula'], 
            telefono=data.get('telefono'),
            direccion=data.get('direccion')
        )
        db.session.add(cli)
        db.session.commit()
        
        # 👥 ¡AQUÍ ADENTRO! Grabamos el registro del nuevo cliente
        registrar_auditoria('CREAR', 'CLIENTES', f'Registrado nuevo cliente: {cli.nombre} (Cédula/RIF: {cli.cedula}).')
        
        return jsonify({'id': cli.id}), 201

@api.route('/api/clientes/<int:id>', methods=['PUT', 'DELETE'])
def modificar_cliente(id):
    cli = Cliente.query.get_or_404(id)
    if request.method == 'PUT':
        data = request.get_json()
        cli.nombre = data.get('nombre', cli.nombre)
        cli.telefono = data.get('telefono', cli.telefono)
        cli.direccion = data.get('direccion', cli.direccion)
        db.session.commit()
        
        registrar_auditoria('EDITAR', 'CLIENTES', f'Modificados datos del cliente: "{cli.nombre}" (ID: {cli.id}).')
        
        return jsonify({'mensaje': 'Cliente actualizado'})
        
    elif request.method == 'DELETE':
        nombre_temporal = cli.nombre 
        db.session.delete(cli)
        db.session.commit()
        
        registrar_auditoria('ELIMINAR', 'CLIENTES', f'Eliminado permanentemente el cliente: "{nombre_temporal}" (ID: {id}).')
        
        return jsonify({'mensaje': 'Cliente eliminado'})

# ---------- API: Ventas ----------
@api.route('/api/ventas', methods=['POST'])
def crear_venta():
    data = request.get_json()
    tipo = data['tipo']
    cliente_id = data.get('cliente_id')
    productos_vendidos = data['productos']

    tasa = 0.0
    ultima_tasa = TasaBCV.query.order_by(TasaBCV.id.desc()).first()
    if not ultima_tasa:
        return jsonify({'error': 'No hay tasa BCV registrada. Ejecute primero la aplicación.'}), 500
    tasa = ultima_tasa.tasa

    total_usd = 0.0
    total_bs = 0.0
    detalles = []
    for item in productos_vendidos:
        prod = Producto.query.get(item['id'])
        if not prod:
            return jsonify({'error': f'Producto con id {item["id"]} no encontrado'}), 400
        cantidad = float(item['cantidad'])
        if prod.stock < cantidad:
            return jsonify({'error': f'Stock insuficiente de {prod.descripcion}'}), 400

        precio_usd = prod.precio
        subtotal_usd = cantidad * precio_usd
        subtotal_bs = subtotal_usd * tasa

        total_usd += subtotal_usd
        total_bs += subtotal_bs
        detalles.append({
            'producto': prod,
            'cantidad': cantidad,
            'precio_unit': precio_usd
        })

    if tipo == 'credito' and not cliente_id:
        return jsonify({'error': 'Debe seleccionar un cliente para venta a crédito'}), 400

    venta = Venta(
        tipo=tipo,
        cliente_id=cliente_id,
        total_bs=total_bs,
        total_usd=total_usd,
        tasa_bcv=tasa
    )
    db.session.add(venta)
    db.session.flush()

    for det in detalles:
        det_venta = DetalleVenta(
            venta_id=venta.id,
            producto_id=det['producto'].id,
            cantidad=det['cantidad'],
            precio_unitario=det['precio_unit']
        )
        db.session.add(det_venta)
        det['producto'].stock = round(det['producto'].stock - det['cantidad'], 2)

    if tipo == 'credito':
        fecha_venc = data.get('fecha_vencimiento')
        if fecha_venc:
            fecha_venc = datetime.strptime(fecha_venc, '%Y-%m-%d').date()
        else:
            fecha_venc = date.today() + timedelta(days=30)
        credito = Credito(
            venta_id=venta.id,
            cliente_id=cliente_id,
            monto_total=total_bs,
            saldo_restante=total_bs,
            fecha_vencimiento=fecha_venc
        )
        db.session.add(credito)
        cliente = Cliente.query.get(cliente_id)
        if cliente:
            cliente.saldo_pendiente += total_bs

    db.session.commit()
    
    # 💰 ¡AQUÍ ADENTRO! Registramos la venta realizada
    registrar_auditoria('CREAR', 'CAJA', f'Venta #{venta.id} procesada a {tipo}. Total: ${total_usd:.2f} USD / Bs.{total_bs:.2f}.')

    return jsonify({
        'mensaje': 'Venta registrada',
        'venta_id': venta.id,
        'total_usd': total_usd,
        'total_bs': total_bs,
        'tasa': tasa
    }), 201

# ---------- API: Créditos ----------
@api.route('/api/creditos', methods=['GET'])
def listar_creditos():
    estado = request.args.get('estado', 'pendiente')
    query = Credito.query
    if estado != 'todos':
        query = query.filter_by(estado=estado)
    creditos = query.order_by(Credito.fecha_vencimiento).all()
    resultado = []
    for c in creditos:
        cliente_nombre = ''
        if c.cliente_id:
            cliente = Cliente.query.get(c.cliente_id)
            if cliente:
                cliente_nombre = cliente.nombre
        resultado.append({
            'id': c.id,
            'venta_id': c.venta_id,
            'cliente_nombre': cliente_nombre,
            'monto_total': c.monto_total,
            'saldo_restante': c.saldo_restante,
            'fecha_vencimiento': c.fecha_vencimiento.isoformat() if c.fecha_vencimiento else '',
            'estado': c.estado,
            'dias_restantes': (c.fecha_vencimiento - date.today()).days if c.fecha_vencimiento else None
        })
    return jsonify(resultado)

@api.route('/api/creditos/<int:id>/abonar', methods=['POST'])
def abonar_credito(id):
    credito = Credito.query.get_or_404(id)
    data = request.get_json()
    monto = round(float(data['monto']), 2)

    if monto <= 0:
        return jsonify({'error': 'Monto debe ser positivo'}), 400
    if monto > credito.saldo_restante + 0.001:
        return jsonify({'error': 'El abono supera el saldo pendiente'}), 400

    abono = Abono(credito_id=id, monto=monto)
    db.session.add(abono)
    credito.saldo_restante = round(credito.saldo_restante - monto, 2)
    if credito.saldo_restante <= 0:
        credito.saldo_restante = 0.0
        credito.estado = 'pagado'

    cliente = Cliente.query.get(credito.cliente_id)
    if cliente:
        cliente.saldo_pendiente = round(cliente.saldo_pendiente - monto, 2)
        if cliente.saldo_pendiente < 0:
            cliente.saldo_pendiente = 0.0

    db.session.commit()
    return jsonify({'mensaje': 'Abono registrado', 'saldo_restante': credito.saldo_restante})

# ---------- Endpoint para contar productos con stock bajo ----------
@api.route('/api/productos/stock-bajo', methods=['GET'])
def contar_stock_bajo():
    cantidad = Producto.query.filter(Producto.stock <= Producto.stock_minimo).count()
    return jsonify({'count': cantidad})

# ---------- COMPRAS / ENTRADAS DE INVENTARIO ----------
@api.route('/api/compras', methods=['GET', 'POST'])
def manejar_compras():
    if request.method == 'GET':
        compras = Compra.query.order_by(Compra.fecha.desc()).all()
        resultado = []
        for c in compras:
            proveedor_nombre = ''
            if c.proveedor_id:
                prov = Proveedor.query.get(c.proveedor_id)
                if prov:
                    proveedor_nombre = prov.nombre
            resultado.append({
                'id': c.id,
                'fecha': c.fecha.strftime('%Y-%m-%d %H:%M'),
                'proveedor': proveedor_nombre,
                'total': c.total,
                'observaciones': c.observaciones
            })
        return jsonify(resultado)
    elif request.method == 'POST':
        data = request.get_json()
        proveedor_id = data.get('proveedor_id')
        observaciones = data.get('observaciones', '')
        productos_comprados = data['productos']

        total_compra = 0.0
        detalles = []
        for item in productos_comprados:
            prod = Producto.query.get(item['id'])
            if not prod:
                return jsonify({'error': f'Producto con id {item["id"]} no encontrado'}), 400
            cantidad = float(item['cantidad'])
            costo_unitario = float(item.get('costo_unitario', 0))
            subtotal = cantidad * costo_unitario
            total_compra += subtotal
            detalles.append({
                'producto': prod,
                'cantidad': cantidad,
                'costo_unitario': costo_unitario
            })

        tipo = data.get('tipo', 'contado')
        compra = Compra(
            proveedor_id=proveedor_id,
            total=total_compra,
            observaciones=observaciones
        )
        db.session.add(compra)
        db.session.flush()

        if tipo == 'credito':
            fecha_venc = data.get('fecha_vencimiento')
            if fecha_venc:
                fecha_venc = datetime.strptime(fecha_venc, '%Y-%m-%d').date()
            else:
                fecha_venc = date.today() + timedelta(days=30)

            credito = CompraCredito(
                compra_id=compra.id,
                proveedor_id=proveedor_id,
                monto_total=total_compra,
                saldo_restante=total_compra,
                fecha_vencimiento=fecha_venc
            )
            db.session.add(credito)
            proveedor = Proveedor.query.get(proveedor_id)
            if proveedor:
                proveedor.saldo_pagar = round(proveedor.saldo_pagar + total_compra, 2)

        for det in detalles:
            det_compra = DetalleCompra(
                compra_id=compra.id,
                producto_id=det['producto'].id,
                cantidad=det['cantidad'],
                costo_unitario=det['costo_unitario']
            )
            db.session.add(det_compra)
            det['producto'].stock += det['cantidad']
            det['producto'].costo = det['costo_unitario']

        db.session.commit()
        
        registrar_auditoria('CREAR', 'COMPRAS', f'Compra de mercancía #{compra.id} registrada. Total invertido: ${total_compra:.2f} USD.')
        
        return jsonify({'mensaje': 'Compra registrada', 'compra_id': compra.id}), 201

@api.route('/api/compras/<int:id>', methods=['GET'])
def detalle_compra(id):
    compra = Compra.query.get_or_404(id)
    proveedor_nombre = ''
    if compra.proveedor_id:
        prov = Proveedor.query.get(compra.proveedor_id)
        if prov:
            proveedor_nombre = prov.nombre
    detalles = []
    for det in compra.detalles:
        prod = Producto.query.get(det.producto_id)
        detalles.append({
            'producto': prod.descripcion if prod else 'Producto eliminado',
            'codigo': prod.codigo if prod else '',
            'cantidad': det.cantidad,
            'costo_unitario': det.costo_unitario,
            'subtotal': det.cantidad * det.costo_unitario
        })
    return jsonify({
        'id': compra.id,
        'fecha': compra.fecha.strftime('%Y-%m-%d %H:%M'),
        'proveedor': proveedor_nombre,
        'total': compra.total,
        'observaciones': compra.observaciones,
        'detalles': detalles
    })

# ---------- Reporte diario de ventas ----------
@api.route('/api/ventas/hoy', methods=['GET'])
def ventas_hoy():
    hoy = date.today()
    inicio = datetime(hoy.year, hoy.month, hoy.day, 0, 0, 0)
    fin = datetime(hoy.year, hoy.month, hoy.day, 23, 59, 59)

    ventas_dia = Venta.query.filter(Venta.fecha >= inicio, Venta.fecha <= fin).order_by(Venta.fecha.desc()).all()

    ventas = []
    for v in ventas_dia:
        if v.tipo == 'contado':
            ventas.append(v)
        elif v.tipo == 'credito':
            credito = Credito.query.filter_by(venta_id=v.id).first()
            if credito and credito.estado == 'pagado':
                ventas.append(v)

    total_usd = 0.0
    total_bs = 0.0
    lista_ventas = []

    for v in ventas:
        total_usd += v.total_usd
        total_bs += v.total_bs

        cliente_nombre = None
        if v.cliente_id:
            cliente = Cliente.query.get(v.cliente_id)
            if cliente:
                cliente_nombre = cliente.nombre

        productos = []
        for det in v.detalles:
            prod = Producto.query.get(det.producto_id)
            productos.append({
                'descripcion': prod.descripcion if prod else 'Producto eliminado',
                'cantidad': det.cantidad
            })

        lista_ventas.append({
            'id': v.id,
            'hora': v.fecha.strftime('%H:%M:%S'),
            'tipo': v.tipo,
            'cliente': cliente_nombre,
            'total_usd': v.total_usd,
            'total_bs': v.total_bs,
            'tasa_bcv': v.tasa_bcv,
            'productos': productos
        })

    abonos_dia = Abono.query.filter(Abono.fecha >= inicio, Abono.fecha <= fin).order_by(Abono.fecha.desc()).all()
    total_abonos_bs = 0.0
    lista_abonos = []
    for a in abonos_dia:
        total_abonos_bs += a.monto
        cliente_nombre = ''
        if a.credito and a.credito.cliente_id:
            cliente = Cliente.query.get(a.credito.cliente_id)
            if cliente:
                cliente_nombre = cliente.nombre
        lista_abonos.append({
            'id': a.id,
            'hora': a.fecha.strftime('%H:%M:%S'),
            'cliente': cliente_nombre,
            'monto': a.monto
        })

    devoluciones_dia = Devolucion.query.filter(Devolucion.fecha >= inicio, Devolucion.fecha <= fin).all()
    total_dev_usd = sum(d.total_usd for d in devoluciones_dia)
    total_dev_bs = sum(d.total_bs for d in devoluciones_dia)

    return jsonify({
        'fecha': hoy.isoformat(),
        'resumen': {
            'cantidad_ventas': len(ventas),
            'total_ventas_usd': total_usd - total_dev_usd,
            'total_ventas_bs': total_bs - total_dev_bs,
            'total_abonos_bs': total_abonos_bs,
            'total_devoluciones_usd': total_dev_usd,
            'total_devoluciones_bs': total_dev_bs,
            'cantidad_abonos': len(abonos_dia)
        },
        'ventas': lista_ventas,
        'abonos': lista_abonos
    })

# ---------- Funciones Excel ----------
@api.route('/api/inventario/exportar')
@login_required
@role_required('admin', 'gerente')
def exportar_inventario():
    productos = Producto.query.all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario"
    ws.append(['Código', 'Descripción', 'Proveedor', 'Costo', 'Precio', 'Stock', 'Stock Mínimo', 'Ubicación'])
    for p in productos:
        ws.append([
            p.codigo,
            p.descripcion,
            p.proveedor.nombre if p.proveedor else '',
            p.costo,
            p.precio,
            p.stock,
            p.stock_minimo,
            p.ubicacion
        ])
    archivo = BytesIO()
    wb.save(archivo)
    archivo.seek(0)
    return send_file(archivo, download_name='inventario.xlsx', as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@api.route('/api/inventario/importar', methods=['POST'])
@login_required
@role_required('admin', 'gerente')
def importar_inventario():
    if 'file' not in request.files:
        return jsonify({'error': 'No se envió archivo'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Archivo sin nombre'}), 400

    filename = file.filename
    ruta = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)
    file.save(ruta)

    try:
        wb = openpyxl.load_workbook(ruta)
        ws = wb.active
        filas_importadas = 0
        errores = []

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if len(row) < 2:
                continue

            codigo = str(row[0]).strip() if row[0] else None
            if not codigo:
                continue

            descripcion = str(row[1]) if row[1] else ''
            proveedor_nombre = str(row[2]) if len(row) > 2 and row[2] else ''

            def a_numero(valor, nombre_columna, fila_num):
                if valor is None or str(valor).strip() == '':
                    return 0.0
                try:
                    return float(valor)
                except ValueError:
                    errores.append(f"Fila {fila_num}, columna '{nombre_columna}': '{valor}' no es un número. Se asumió 0.")
                    return 0.0

            costo = a_numero(row[3] if len(row) > 3 else None, 'Costo', idx)
            precio = a_numero(row[4] if len(row) > 4 else None, 'Precio', idx)
            stock = a_numero(row[5] if len(row) > 5 else None, 'Stock', idx)
            stock_min = a_numero(row[6] if len(row) > 6 else None, 'Stock Mínimo', idx)
            ubicacion = str(row[7]) if len(row) > 7 and row[7] else ''

            proveedor_id = None
            if proveedor_nombre:
                prov = Proveedor.query.filter_by(nombre=proveedor_nombre).first()
                if not prov:
                    prov = Proveedor(nombre=proveedor_nombre)
                    db.session.add(prov)
                    db.session.flush()
                proveedor_id = prov.id

            producto = Producto.query.filter_by(codigo=codigo).first()
            if producto:
                producto.descripcion = descripcion
                producto.proveedor_id = proveedor_id or producto.proveedor_id
                producto.costo = costo
                producto.precio = precio
                producto.stock = stock
                producto.stock_minimo = stock_min
                producto.ubicacion = ubicacion
            else:
                nuevo = Producto(
                    codigo=codigo,
                    descripcion=descripcion,
                    proveedor_id=proveedor_id,
                    costo=costo,
                    precio=precio,
                    stock=stock,
                    stock_minimo=stock_min,
                    ubicacion=ubicacion
                )
                db.session.add(nuevo)
            filas_importadas += 1

        db.session.commit()
        os.remove(ruta)

        mensaje = f'Se importaron/actualizaron {filas_importadas} productos.'
        if errores:
            mensaje += f' Se encontraron {len(errores)} advertencias (revisa la consola o los logs).'
            for e in errores:
                print(e)
        return jsonify({'mensaje': mensaje})
    except Exception as e:
        return jsonify({'error': f'Error al procesar el archivo: {str(e)}'}), 500

@api.route('/api/respaldo/completo')
@login_required
@role_required('admin')
def respaldo_completo():
    wb = openpyxl.Workbook()
    # Productos
    ws1 = wb.active
    ws1.title = "Productos"
    ws1.append(['ID', 'Código', 'Descripción', 'Proveedor', 'Costo', 'Precio', 'Stock'])
    for p in Producto.query.all():
        ws1.append([p.id, p.codigo, p.descripcion,
                    p.proveedor.nombre if p.proveedor else '', p.costo, p.precio, p.stock])

    # Clientes
    ws2 = wb.create_sheet("Clientes")
    ws2.append(['ID', 'Nombre', 'Cédula', 'Teléfono', 'Saldo Pendiente'])
    for c in Cliente.query.all():
        ws2.append([c.id, c.nombre, c.cedula, c.telefono, c.saldo_pendiente])

    # Ventas
    ws3 = wb.create_sheet("Ventas")
    ws3.append(['ID', 'Fecha', 'Tipo', 'Cliente', 'Total USD', 'Total Bs', 'Tasa BCV'])
    for v in Venta.query.all():
        cliente_nombre = 'Mostrador'
        if v.cliente_id:
            cliente = Cliente.query.get(v.cliente_id)
            if cliente:
                cliente_nombre = cliente.nombre
        ws3.append([v.id, v.fecha.strftime('%Y-%m-%d %H:%M'), v.tipo, cliente_nombre,
                    v.total_usd, v.total_bs, v.tasa_bcv])

    # Créditos Clientes
    ws4 = wb.create_sheet("Creditos Clientes")
    ws4.append(['ID', 'Cliente', 'Monto Total', 'Saldo Restante', 'Vencimiento', 'Estado'])
    for c in Credito.query.all():
        cliente_nombre = ''
        if c.cliente_id:
            cliente = Cliente.query.get(c.cliente_id)
            if cliente:
                cliente_nombre = cliente.nombre
        ws4.append([c.id, cliente_nombre, c.monto_total, c.saldo_restante,
                    c.fecha_vencimiento.isoformat() if c.fecha_vencimiento else '', c.estado])

    archivo = BytesIO()
    wb.save(archivo)
    archivo.seek(0)
    return send_file(archivo, download_name='respaldo_completo.xlsx', as_attachment=True)

# ---------- CRÉDITOS DE PROVEEDORES ----------
@api.route('/api/compras/creditos', methods=['GET'])
def listar_creditos_compras():
    estado = request.args.get('estado', 'pendiente')
    query = CompraCredito.query
    if estado != 'todos':
        query = query.filter_by(estado=estado)
    creditos = query.order_by(CompraCredito.fecha_vencimiento).all()
    resultado = []
    for c in creditos:
        proveedor_nombre = ''
        prov = Proveedor.query.get(c.proveedor_id)
        if prov:
            proveedor_nombre = prov.nombre
        resultado.append({
            'id': c.id,
            'compra_id': c.compra_id,
            'proveedor_nombre': proveedor_nombre,
            'monto_total': c.monto_total,
            'saldo_restante': c.saldo_restante,
            'fecha_vencimiento': c.fecha_vencimiento.isoformat() if c.fecha_vencimiento else '',
            'estado': c.estado,
            'dias_restantes': (c.fecha_vencimiento - date.today()).days if c.fecha_vencimiento else None
        })
    return jsonify(resultado)

@api.route('/api/compras/creditos/<int:id>/pagar', methods=['POST'])
def pagar_credito_compra(id):
    credito = CompraCredito.query.get_or_404(id)
    data = request.get_json()
    monto = round(float(data['monto']), 2)
    if monto <= 0:
        return jsonify({'error': 'Monto debe ser positivo'}), 400
    if monto > credito.saldo_restante + 0.001:
        return jsonify({'error': 'El pago supera el saldo pendiente'}), 400

    pago = PagoProveedor(compra_credito_id=id, monto=monto)
    db.session.add(pago)
    credito.saldo_restante = round(credito.saldo_restante - monto, 2)
    if credito.saldo_restante <= 0:
        credito.saldo_restante = 0.0
        credito.estado = 'pagado'

    proveedor = Proveedor.query.get(credito.proveedor_id)
    if proveedor:
        proveedor.saldo_pagar = round(proveedor.saldo_pagar - monto, 2)
        if proveedor.saldo_pagar < 0:
            proveedor.saldo_pagar = 0.0

    db.session.commit()
    return jsonify({'mensaje': 'Pago registrado', 'saldo_restante': credito.saldo_restante})

# ---------- DEUDAS GENERALES (gastos) ----------
@api.route('/api/deudas', methods=['GET', 'POST'])
def manejar_deudas():
    if request.method == 'GET':
        deudas = DeudaGeneral.query.order_by(DeudaGeneral.fecha_vencimiento).all()
        return jsonify([{
            'id': d.id,
            'descripcion': d.descripcion,
            'monto_total': d.monto_total,
            'saldo_restante': d.saldo_restante,
            'fecha_vencimiento': d.fecha_vencimiento.isoformat() if d.fecha_vencimiento else '',
            'estado': d.estado,
            'dias_restantes': (d.fecha_vencimiento - date.today()).days if d.fecha_vencimiento else None
        } for d in deudas])
    elif request.method == 'POST':
        data = request.get_json()
        fecha_venc = data.get('fecha_vencimiento')
        if fecha_venc:
            fecha_venc = datetime.strptime(fecha_venc, '%Y-%m-%d').date()
        else:
            fecha_venc = date.today() + timedelta(days=30)
        deuda = DeudaGeneral(
            descripcion=data['descripcion'],
            monto_total=data['monto_total'],
            saldo_restante=data['monto_total'],
            fecha_vencimiento=fecha_venc
        )
        db.session.add(deuda)
        db.session.commit()
        return jsonify({'mensaje': 'Deuda registrada', 'id': deuda.id}), 201

@api.route('/api/deudas/<int:id>/pagar', methods=['POST'])
def pagar_deuda(id):
    deuda = DeudaGeneral.query.get_or_404(id)
    data = request.get_json()
    monto = round(float(data['monto']), 2)
    if monto <= 0:
        return jsonify({'error': 'Monto debe ser positivo'}), 400
    if monto > deuda.saldo_restante + 0.001:
        return jsonify({'error': 'El pago supera el saldo pendiente'}), 400

    pago = PagoDeudaGeneral(deuda_id=id, monto=monto)
    db.session.add(pago)
    deuda.saldo_restante = round(deuda.saldo_restante - monto, 2)
    if deuda.saldo_restante <= 0:
        deuda.saldo_restante = 0.0
        deuda.estado = 'pagado'
    db.session.commit()
    return jsonify({'mensaje': 'Pago registrado', 'saldo_restante': deuda.saldo_restante})

# ---------- Búsqueda rápida de productos ----------
@api.route('/api/productos/buscar', methods=['GET'])
def buscar_productos():
    q = request.args.get('q', '').strip()
    if not q:
        return jsonify([])
    productos = Producto.query.filter(
        db.or_(
            Producto.codigo.ilike(f'%{q}%'),
            Producto.descripcion.ilike(f'%{q}%')
        )
    ).limit(10).all()
    return jsonify([{
        'id': p.id,
        'codigo': p.codigo,
        'descripcion': p.descripcion,
        'precio': p.precio,
        'stock': p.stock
    } for p in productos])

# ---------- DEVOLUCIONES ----------
@api.route('/api/ventas/buscar', methods=['GET'])
def buscar_venta_por_id():
    venta_id = request.args.get('id', type=int)
    if not venta_id:
        return jsonify({'error': 'Debe proporcionar un ID de venta'}), 400
    venta = Venta.query.get(venta_id)
    if not venta:
        return jsonify({'error': 'Venta no encontrada'}), 404
    cliente_nombre = 'Mostrador'
    if venta.cliente_id:
        cliente = Cliente.query.get(venta.cliente_id)
        if cliente:
            cliente_nombre = cliente.nombre
    productos = []
    for det in venta.detalles:
        prod = Producto.query.get(det.producto_id)
        if prod:
            productos.append({
                'id': prod.id,
                'codigo': prod.codigo,
                'descripcion': prod.descripcion,
                'cantidad_vendida': det.cantidad,
                'precio_unitario': det.precio_unitario
            })
    return jsonify({
        'id': venta.id,
        'fecha': venta.fecha.strftime('%Y-%m-%d %H:%M'),
        'tipo': venta.tipo,
        'cliente': cliente_nombre,
        'total_bs': venta.total_bs,
        'total_usd': venta.total_usd,
        'tasa_bcv': venta.tasa_bcv,
        'productos': productos
    })

@api.route('/api/devoluciones', methods=['POST'])
def crear_devolucion():
    data = request.get_json()
    venta_id = data['venta_id']
    productos_devueltos = data['productos']
    observaciones = data.get('observaciones', '')

    venta = Venta.query.get(venta_id)
    if not venta:
        return jsonify({'error': 'Venta no encontrada'}), 404

    total_usd = 0.0
    total_bs = 0.0
    for item in productos_devueltos:
        prod_id = item['id']
        cantidad_dev = float(item['cantidad'])
        det_venta = DetalleVenta.query.filter_by(venta_id=venta_id, producto_id=prod_id).first()
        if not det_venta:
            return jsonify({'error': f'Producto {prod_id} no pertenece a la venta'}), 400
        if cantidad_dev > det_venta.cantidad:
            return jsonify({'error': f'Cantidad a devolver de {prod_id} supera la vendida ({det_venta.cantidad})'}), 400
        precio_usd = det_venta.precio_unitario
        subtotal_usd = cantidad_dev * precio_usd
        total_usd += subtotal_usd

    total_bs = total_usd * venta.tasa_bcv

    devolucion = Devolucion(
        venta_id=venta_id,
        total_usd=total_usd,
        total_bs=total_bs,
        observaciones=observaciones
    )
    db.session.add(devolucion)
    db.session.flush()

    for item in productos_devueltos:
        prod_id = item['id']
        cantidad_dev = float(item['cantidad'])
        prod = Producto.query.get(prod_id)
        det_venta = DetalleVenta.query.filter_by(venta_id=venta_id, producto_id=prod_id).first()
        det_dev = DetalleDevolucion(
            devolucion_id=devolucion.id,
            producto_id=prod_id,
            cantidad=cantidad_dev,
            precio_unitario=det_venta.precio_unitario
        )
        db.session.add(det_dev)
        prod.stock = round(prod.stock + cantidad_dev, 2)

    if venta.tipo == 'credito':
        credito = Credito.query.filter_by(venta_id=venta_id).first()
        if credito:
            credito.monto_total = round(credito.monto_total - total_bs, 2)
            credito.saldo_restante = round(credito.saldo_restante - total_bs, 2)
            if credito.saldo_restante <= 0:
                credito.saldo_restante = 0.0
                credito.estado = 'pagado'
            cliente = Cliente.query.get(venta.cliente_id)
            if cliente:
                cliente.saldo_pendiente = round(cliente.saldo_pendiente - total_bs, 2)
                if cliente.saldo_pendiente < 0:
                    cliente.saldo_pendiente = 0.0

    db.session.commit()
    
    registrar_auditoria('DEVOLUCION', 'DEVOLUCIONES', f'Devolución #{devolucion.id} registrada para la Venta #{venta_id}. Monto devuelto: ${total_usd:.2f} USD.')
    
    return jsonify({'mensaje': 'Devolución registrada', 'id': devolucion.id, 'total_usd': total_usd, 'total_bs': total_bs}), 201

@api.route('/api/devoluciones/hoy', methods=['GET'])
def listar_devoluciones_hoy():
    hoy = date.today()
    inicio = datetime(hoy.year, hoy.month, hoy.day, 0, 0, 0)
    fin = datetime(hoy.year, hoy.month, hoy.day, 23, 59, 59)
    devoluciones = Devolucion.query.filter(Devolucion.fecha >= inicio, Devolucion.fecha <= fin).order_by(Devolucion.fecha.desc()).all()
    resultado = []
    for d in devoluciones:
        cliente_nombre = ''
        if d.venta and d.venta.cliente_id:
            cliente = Cliente.query.get(d.venta.cliente_id)
            if cliente:
                cliente_nombre = cliente.nombre
        productos = []
        for det in d.detalles:
            prod = Producto.query.get(det.producto_id)
            productos.append({
                'descripcion': prod.descripcion if prod else 'Eliminado',
                'cantidad': det.cantidad,
                'precio_unitario': det.precio_unitario
            })
        resultado.append({
            'id': d.id,
            'fecha': d.fecha.strftime('%H:%M:%S'),
            'cliente': cliente_nombre,
            'total_usd': d.total_usd,
            'total_bs': d.total_bs,
            'observaciones': d.observaciones,
            'productos': productos
        })
    return jsonify(resultado)

@api.route('/api/ventas/todas', methods=['GET'])
def todas_las_ventas():
    ventas = Venta.query.order_by(Venta.id.desc()).all()
    resultado = []
    for v in ventas:
        cliente_nombre = 'Mostrador'
        if v.cliente_id:
            cliente = Cliente.query.get(v.cliente_id)
            if cliente:
                cliente_nombre = cliente.nombre

        if v.tipo == 'contado':
            estado = 'Pagada'
        else:
            credito = Credito.query.filter_by(venta_id=v.id).first()
            if credito and credito.estado == 'pagado':
                estado = 'Pagada'
            else:
                estado = 'Pendiente'

        devolucion = Devolucion.query.filter_by(venta_id=v.id).first()
        monto_dev = devolucion.total_usd if devolucion else 0.0

        resultado.append({
            'id': v.id,
            'fecha': v.fecha.strftime('%Y-%m-%d %H:%M'),
            'cliente': cliente_nombre,
            'tipo': v.tipo,
            'total_usd': v.total_usd,
            'total_bs': v.total_bs,
            'estado': estado,
            'devolucion_usd': monto_dev
        })

    return jsonify(resultado)

@api.route('/api/ventas/por-cliente', methods=['GET'])
def ventas_por_cliente():
    cedula = request.args.get('cedula', '').strip()
    if not cedula:
        return jsonify([])

    cliente = Cliente.query.filter_by(cedula=cedula).first()
    if not cliente:
        return jsonify([])

    ventas = Venta.query.filter_by(cliente_id=cliente.id) \
                        .order_by(Venta.id.desc()) \
                        .limit(50) \
                        .all()

    resultado = []
    for v in ventas:
        estado = 'Pagada' if v.tipo == 'contado' else 'Pendiente'
        if v.tipo == 'credito':
            credito = Credito.query.filter_by(venta_id=v.id).first()
            if credito and credito.estado == 'pagado':
                estado = 'Pagada'

        devolucion = Devolucion.query.filter_by(venta_id=v.id).first()
        devuelto = devolucion.total_usd if devolucion else 0.0

        resultado.append({
            'id': v.id,
            'fecha': v.fecha.strftime('%Y-%m-%d %H:%M'),
            'tipo': v.tipo,
            'total_usd': v.total_usd,
            'total_bs': v.total_bs,
            'estado': estado,
            'devuelto_usd': devuelto
        })

    return jsonify({
        'cliente': cliente.nombre,
        'cedula': cliente.cedula,
        'ventas': resultado
    })